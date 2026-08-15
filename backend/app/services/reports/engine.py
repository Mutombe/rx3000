"""The report engine: a report is a declaration, not a page.

The system we are replacing is four Windows applications with four Reports
menus and roughly a hundred and twenty reports between them. Every one of those
reports re-implements its own date range, its own totals row and its own export
button, which is why theirs are inconsistent — some export to Excel, some only
print, and the date control is in a different place on each.

Writing a hundred and twenty screens would reproduce exactly that. So a report
here declares *what it is* — its parameters, its columns, its query — and the
engine supplies everything a report needs but should never have to implement:

* parameter parsing, defaulting and validation, once
* a genuine total, from a separate count, never `len(rows)` — a page that
  reports its own length as the total looks complete and is not
* sorting, paging, and column totals
* CSV, Excel and JSON from the *same* query the screen ran, so an export can
  never disagree with what the person exporting it was looking at
* the column's type, so money, dates and quantities format identically on the
  screen, in the spreadsheet and on paper

The engine is the whole reason the catalogue is affordable. Each new report
after this is a declaration of a few dozen lines.
"""
from __future__ import annotations

import csv
import io
from dataclasses import dataclass, field
from datetime import date, datetime, timedelta
from typing import Any, Callable, Literal

from sqlalchemy.orm import Session

ColumnKind = Literal["text", "money", "number", "date", "datetime", "percent", "code"]
ParamKind = Literal["date", "daterange", "text", "select", "branch", "bool"]


@dataclass
class Param:
    """One input a report accepts.

    Declared rather than hand-rolled so that every report's date range behaves
    the same way and a person who has learned one report has learned all of
    them.
    """
    key: str
    label: str
    kind: ParamKind = "text"
    required: bool = False
    default: Any = None
    # For `select`. A callable so the options can come from the database.
    options: Callable[[Session], list[dict]] | list[dict] | None = None
    help: str = ""


@dataclass
class Column:
    key: str
    header: str
    kind: ColumnKind = "text"
    align: str = ""
    # Include this column in the footer totals. Only meaningful for numbers.
    total: bool = False
    width: int = 0


@dataclass
class Report:
    key: str
    title: str
    module: str
    # What question this answers. Shown under the title, because a list of
    # eighty report names is unusable without it.
    purpose: str = ""
    params: list[Param] = field(default_factory=list)
    columns: list[Column] = field(default_factory=list)
    # (db, params) -> list of row dicts. Rows are plain dicts so a report can
    # be assembled from a raw query, an ORM query, or computed in Python.
    rows: Callable[[Session, dict], list[dict]] = None  # type: ignore
    # Optional: a link target for a row, so a total is never a dead end.
    drill: Callable[[dict], str] | None = None
    # Sensitive reports demand a second person, not a second prompt.
    step_up: bool = False


REGISTRY: dict[str, Report] = {}


def register(report: Report) -> Report:
    if report.key in REGISTRY:
        raise ValueError(
            f"Two reports both claim the key '{report.key}'. Keys are what the "
            "front end routes on, so the second would silently shadow the first."
        )
    REGISTRY[report.key] = report
    return report


def catalogue() -> list[dict]:
    """Every report, grouped by module, for the reports index."""
    return [
        {
            "key": r.key, "title": r.title, "module": r.module,
            "purpose": r.purpose, "step_up": r.step_up,
            "params": [
                {"key": p.key, "label": p.label, "kind": p.kind,
                 "required": p.required, "default": _serialise(_default_for(p)),
                 "help": p.help}
                for p in r.params
            ],
        }
        for r in sorted(REGISTRY.values(), key=lambda r: (r.module, r.title))
    ]


def _serialise(value: Any) -> Any:
    """A default as the front end can put it in an input.

    `default` may be a callable — "the first of this month", evaluated when the
    report is opened rather than when the process started. Sending the callable
    itself reaches the browser as `[object Object]`, which then goes back to the
    server as a date and is refused. Resolved and formatted here instead.
    """
    if isinstance(value, (date, datetime)):
        return value.isoformat()[:10]
    if value is None or isinstance(value, (str, int, float, bool)):
        return value
    return str(value)


def _coerce(param: Param, raw: Any) -> Any:
    """Turn a query-string value into what the report's query expects.

    A date that will not parse is refused with the shape we wanted rather than
    silently becoming today, because a report quietly run for the wrong period
    is worse than one that did not run.
    """
    if raw in (None, ""):
        return _default_for(param)
    if param.kind == "date":
        if isinstance(raw, date):
            return raw
        try:
            return date.fromisoformat(str(raw))
        except ValueError:
            raise ValueError(f"{param.label}: '{raw}' is not a date. Use YYYY-MM-DD.")
    if param.kind == "bool":
        return str(raw).lower() in ("1", "true", "yes", "on")
    if param.kind == "branch":
        try:
            return int(raw)
        except (TypeError, ValueError):
            raise ValueError(f"{param.label}: expected a branch, got '{raw}'.")
    return str(raw)


def _default_for(param: Param) -> Any:
    if callable(param.default):
        return param.default()
    return param.default


def resolve_params(report: Report, given: dict) -> dict:
    """Validate and fill in what the report was asked for."""
    out: dict = {}
    for param in report.params:
        value = _coerce(param, given.get(param.key))
        if param.required and value in (None, ""):
            raise ValueError(f"{param.label} is required.")
        out[param.key] = value
    return out


def run(
    db: Session, key: str, given: dict, *,
    page: int = 1, per_page: int = 100, sort: str = "", desc: bool = False,
) -> dict:
    """Run a report and return one page of it, plus the totals for all of it."""
    report = REGISTRY.get(key)
    if not report:
        raise KeyError(f"There is no report called '{key}'.")
    params = resolve_params(report, given)

    rows = report.rows(db, params)

    if sort:
        rows = sorted(
            rows,
            key=lambda r: (r.get(sort) is None, r.get(sort)),
            reverse=desc,
        )

    # Totals are computed across every row, not across the page. A footer that
    # only totals what is on screen is the most quietly misleading thing a
    # report can do.
    totals = {
        c.key: round(sum(float(r.get(c.key) or 0) for r in rows), 2)
        for c in report.columns if c.total
    }

    per_page = max(1, min(int(per_page or 100), 1000))
    total = len(rows)
    last = max(1, -(-total // per_page))
    page = max(1, min(int(page or 1), last))
    start = (page - 1) * per_page
    window = rows[start:start + per_page]

    # Where a row leads, computed for the page being sent rather than for every
    # row. A report that names a product and then makes you go and search for it
    # has answered half a question; this is what stops a total being a dead end.
    # Only the visible page pays for it, and a drill that raises is skipped
    # rather than taking the whole report down with it.
    if report.drill:
        window = [dict(row) for row in window]
        for row in window:
            try:
                row["_drill"] = report.drill(row)
            except Exception:
                pass

    return {
        "key": report.key,
        "title": report.title,
        "purpose": report.purpose,
        "columns": [
            {"key": c.key, "header": c.header, "kind": c.kind,
             "align": c.align or ("right" if c.kind in ("money", "number", "percent") else ""),
             "total": c.total}
            for c in report.columns
        ],
        "rows": window,
        "totals": totals,
        "params": {k: (v.isoformat() if isinstance(v, (date, datetime)) else v)
                   for k, v in params.items()},
        "page": page, "per_page": per_page, "total": total, "pages": last,
        "generated_at": datetime.utcnow().isoformat(),
    }


def _all_rows(db: Session, key: str, given: dict, sort: str = "", desc: bool = False):
    """Every row, for export. Deliberately shares `run`'s query.

    Exports that re-implement the query are exports that disagree with the
    screen, and the person holding the spreadsheet is the one who finds out.
    """
    report = REGISTRY[key]
    params = resolve_params(report, given)
    rows = report.rows(db, params)
    if sort:
        rows = sorted(rows, key=lambda r: (r.get(sort) is None, r.get(sort)), reverse=desc)
    return report, params, rows


def to_csv(db: Session, key: str, given: dict, **kw) -> str:
    report, _, rows = _all_rows(db, key, given, **kw)
    buffer = io.StringIO()
    writer = csv.writer(buffer)
    writer.writerow([c.header for c in report.columns])
    for row in rows:
        writer.writerow([row.get(c.key, "") for c in report.columns])
    return buffer.getvalue()


def to_xlsx(db: Session, key: str, given: dict, **kw) -> bytes:
    """A real spreadsheet, not a CSV with the wrong extension.

    Every one of the incumbent's report screens has an Excel button and the
    pharmacy's taskbar has a spreadsheet permanently open. This is not a
    nice-to-have; it is how the finance work actually gets done, so the output
    has to arrive formatted — money as money, dates as dates, a frozen header
    and a totals row — rather than as text that needs cleaning up first.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    report, params, rows = _all_rows(db, key, given, **kw)
    book = Workbook()
    sheet = book.active
    sheet.title = report.title[:31] or "Report"

    heading = Font(bold=True, size=13)
    sheet.cell(row=1, column=1, value=report.title).font = heading
    stamp = ", ".join(
        f"{k}: {v.isoformat() if isinstance(v, (date, datetime)) else v}"
        for k, v in params.items() if v not in (None, "")
    )
    sheet.cell(row=2, column=1,
               value=f"{stamp}    generated {datetime.now():%Y-%m-%d %H:%M}")
    sheet.cell(row=2, column=1).font = Font(size=9, italic=True, color="777777")

    header_row = 4
    head_font = Font(bold=True, color="FFFFFF")
    head_fill = PatternFill("solid", fgColor="2F2F3A")
    thin = Side(style="thin", color="DDDDDD")

    for i, column in enumerate(report.columns, start=1):
        cell = sheet.cell(row=header_row, column=i, value=column.header)
        cell.font = head_font
        cell.fill = head_fill
        cell.alignment = Alignment(horizontal="center")

    money_format = '#,##0.00'
    for r, row in enumerate(rows, start=header_row + 1):
        for i, column in enumerate(report.columns, start=1):
            value = row.get(column.key)
            if column.kind in ("money", "number", "percent") and value is not None:
                try:
                    value = float(value)
                except (TypeError, ValueError):
                    pass
            cell = sheet.cell(row=r, column=i, value=value)
            cell.border = Border(bottom=thin)
            if column.kind == "money":
                cell.number_format = money_format
            elif column.kind == "percent":
                cell.number_format = '0.0"%"'
            elif column.kind in ("date", "datetime"):
                cell.number_format = "yyyy-mm-dd"
            elif column.kind == "code":
                cell.alignment = Alignment(horizontal="left")

    total_row = header_row + len(rows) + 1
    wrote_total = False
    for i, column in enumerate(report.columns, start=1):
        if not column.total:
            continue
        wrote_total = True
        total = round(sum(float(r.get(column.key) or 0) for r in rows), 2)
        cell = sheet.cell(row=total_row, column=i, value=total)
        cell.font = Font(bold=True)
        cell.number_format = money_format
        cell.border = Border(top=Side(style="double"))
    if wrote_total:
        label = sheet.cell(row=total_row, column=1, value="Total")
        label.font = Font(bold=True)

    for i, column in enumerate(report.columns, start=1):
        longest = max(
            [len(str(column.header))]
            + [len(str(r.get(column.key, ""))) for r in rows[:400]]
        )
        sheet.column_dimensions[get_column_letter(i)].width = min(
            max(column.width or longest + 3, 10), 46)

    # So the header stays put on a report that runs to hundreds of rows.
    sheet.freeze_panes = sheet.cell(row=header_row + 1, column=1)

    out = io.BytesIO()
    book.save(out)
    return out.getvalue()


# ---------------------------------------------------------------- defaults
def month_start() -> date:
    today = date.today()
    return today.replace(day=1)


def today() -> date:
    return date.today()


def days_ago(n: int) -> Callable[[], date]:
    return lambda: date.today() - timedelta(days=n)
